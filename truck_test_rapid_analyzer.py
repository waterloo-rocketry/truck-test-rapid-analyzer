from openpyxl import Workbook as WB
from openpyxl import load_workbook as load_wb

from openpyxl.utils import get_column_letter
from openpyxl.utils import quote_sheetname as quote
from openpyxl.utils import absolute_coordinate as abs_coords

from openpyxl.workbook.defined_name import DefinedName as DName

from openpyxl.chart import LineChart, ScatterChart, Reference, Series

import csv
import os
import sys

from openpyxl.styles.alignment import Alignment

from config_variables import *


def space_columns(target_sheet, depth=1, cols=0):
    '''
    Space the columns to make sure all text in headings is visible.

    Parameters
    ----------
    target_sheet: openpyxl.Worksheet
        The sheet where the columns must be spaced.
    depth: int=1
        How deep the length check should be done, for vertical headers. Default value is 1
    cols: int=0
        How many columns should be checked. Default values is 0, which specifies a scan of all
        columns.
    '''

    exit_flag = False
    col_count = 1

    while not exit_flag:
        max_width = 0
        for didx in range(1, depth + 1):
            if isinstance(target_sheet.cell(didx, col_count).value, str):
                curr_width = len(target_sheet.cell(didx, col_count).value)

                target_sheet.cell(didx, col_count).alignment = Alignment(
                    wrap_text=True)

                if curr_width > max_width:
                    max_width = curr_width

        target_sheet.column_dimensions[get_column_letter(
            col_count)].width = int(0.75 * max_width)
        target_sheet.cell(1, col_count).alignment = Alignment(
            wrap_text=True)

        col_count += 1

        # Exit condition
        if cols:  # If a specific number of columns was specified
            if col_count >= cols:
                exit_flag = True
        else:  # If function is to scan columns until they are empty
            if not target_sheet.cell(1, col_count).value:
                exit_flag = True


class HeaderManager():
    TIME_COL = 1
    ANEMOMETER_RAW_COL = 2
    LOAD_CELL_RAW_COL = 3
    ANEMOMETER_CALIBRATED_COL = 4
    LOAD_CELL_CALIBRATED_COL = 5

    LOAD_CELL_AVERAGED_COL = 6
    ANEMOMETER_AVERAGED_COL = 7
    ANEMOMETER_FT_S_COL = 8

    TARGET_FORCE_COL = 9
    DRAG_AREA_COL = 10
    DRAG_COEFF_COL = 11

    def create_headers(self, sheet, target_row=1):
        """
        Create headers in specified sheet.

        Parameters
        ----------
        sheet: openpyxl.Worksheet
            The sheet where the headers are to be created

        target_row: The row where the headers should be inserted. Default is 1, which is the first row.
                    (openpyxl starts counting at 1 so this is the true first row in the spreadsheet)

        """
        sheet.cell(target_row, self.TIME_COL, "Time (ms)")
        sheet.cell(target_row, self.ANEMOMETER_RAW_COL, "Anemometer Raw (m/s)")
        sheet.cell(target_row, self.LOAD_CELL_RAW_COL, "Load Cell Raw (lbf)")
        sheet.cell(target_row, self.ANEMOMETER_CALIBRATED_COL,
                   "Anemometer Calibrated (m/s)")
        sheet.cell(target_row, self.LOAD_CELL_CALIBRATED_COL,
                   "Load Cell Calibrated (lbf)")

        sheet.cell(target_row, self.LOAD_CELL_AVERAGED_COL,
                   "Load Cell Averaged (lbf)")
        sheet.cell(target_row, self.ANEMOMETER_AVERAGED_COL,
                   "Anemometer Averaged (m/s)")
        sheet.cell(target_row, self.ANEMOMETER_FT_S_COL, "Anemometer (ft/s)")

        sheet.cell(target_row, self.TARGET_FORCE_COL, "Target Force (lbf)")
        sheet.cell(target_row, self.DRAG_AREA_COL, "Drag Area [CdSo] (ft^2)")
        sheet.cell(target_row, self.DRAG_COEFF_COL,
                   "Drag Coefficient (unitless)")


def process_sheet_row(raw_data, row_idx, sheet):
    """
    Create all formulas and input sheet data for a single data timestamp (row)

    Parameters
    ----------
    raw_data: array of float-like str
        The raw data from the DAQ csv
    row_idx: int
        The index of the row where it is the to be inserted. That is, the target excel sheet row
    sheet: openpyxl.Worksheet
        The worksheet where the data will be inputted.
    """
    sheet.cell(row_idx, 1, int(float(raw_data[0])*1000))

    sheet.cell(row_idx, 2, float(raw_data[1]))
    sheet.cell(row_idx, 2).number_format = "0.000"

    sheet.cell(row_idx, 3, float(raw_data[2]))
    sheet.cell(row_idx, 3).number_format = "0.000"

    r_force = "C" + str(row_idx)
    r_wspeed = "B" + str(row_idx)

    calibrated_windspeed_formula = "=" +\
        r_wspeed + "/" + 'ANEMOMETER_FACTOR'
    sheet.cell(row_idx, 4, calibrated_windspeed_formula)
    sheet.cell(row_idx, 4).number_format = "0.000"
    cal_ws = "D" + str(row_idx)

    calibrated_force_formula = "=" + r_force + \
        "/" + 'LOAD_CELL_FACTOR'
    sheet.cell(row_idx, 5, calibrated_force_formula)
    sheet.cell(row_idx, 5).number_format = "0.000"
    cal_force = "E" + str(row_idx)

    force_average_window = 3
    if row_idx - force_average_window >= 1:
        averaged_force_lower_bound = row_idx - force_average_window
    else:
        averaged_force_lower_bound = 1

    averaged_force_upper_bound = row_idx + force_average_window

    f_bounds = cal_force[0] + str(averaged_force_lower_bound) + ":" +\
        cal_force[0] + str(averaged_force_upper_bound)

    averaged_force_formula = "=AVERAGE(" + f_bounds + ")"
    sheet.cell(row_idx, 6, averaged_force_formula)
    sheet.cell(row_idx, 6).number_format = "0.000"
    av_force = "F" + str(row_idx)

    #

    ws_average_window = 0
    if row_idx - ws_average_window >= 1:
        averaged_ws_lower_bound = row_idx - ws_average_window
    else:
        averaged_ws_lower_bound = 1

    averaged_ws_upper_bound = row_idx + ws_average_window

    ws_bounds = cal_ws[0] + str(averaged_ws_lower_bound) + ":" +\
        cal_ws[0] + str(averaged_ws_upper_bound)

    averaged_windspeed_formula = "=AVERAGE(" + ws_bounds + ")"
    sheet.cell(row_idx, 7, averaged_windspeed_formula)
    sheet.cell(row_idx, 7).number_format = "0.000"
    av_windspeed = "G" + str(row_idx)

    ft_s_windspeed_formula = "=" + av_windspeed + "/0.3048"
    sheet.cell(row_idx, 8, ft_s_windspeed_formula)
    sheet.cell(row_idx, 8).number_format = "0.000"
    av_windspeed_ft_s = "H" + str(row_idx)

    target_force_formula = "=(" + av_windspeed_ft_s +\
        "^2)*" + 'AIR_DENSITY_SLG_FT3' +\
        "*" + 'TARGET_DRAG_AREA_FT2' +\
        "*0.5"

    sheet.cell(row_idx, 9, target_force_formula)
    sheet.cell(row_idx, 9).number_format = "0.000"
    target_force_lbf = "I" + str(row_idx)

    drag_area_formula = "=if(" + av_windspeed_ft_s + "=0, ," +\
        "(2*" + av_force + ")/(" + 'AIR_DENSITY_SLG_FT3' +\
        "*(" + av_windspeed_ft_s + ")^2))"
    sheet.cell(row_idx, 10, drag_area_formula)
    sheet.cell(row_idx, 10).number_format = "0.000"
    drag_area_ft2 = "J" + str(row_idx)

    coeff_of_drag_formula = "=" + drag_area_ft2 +\
        "/" + 'NOM_SA_FT2'
    sheet.cell(row_idx, 11, coeff_of_drag_formula)
    sheet.cell(row_idx, 11).number_format = "0.000"
    coeff_of_drag = "K" + str(row_idx)


def create_graphs(data_sheet, output_sheet, max_idx):
    """
    Create graphs from the data.

    The graphs unfortunately do not transfer to google sheets formats, so they will need to be
    re-created once in google sheets.

    Parameters
    ----------
    data_sheet: openpyxl.Worksheet
        The sheet where the data is located
    output_sheet: openpyxl.Worksheet
        The sheet where the graph should be located
    max_idx: int
        The maximum shreadsheet row where the data is
    """

    chart = ScatterChart()
    drag_area_chart = ScatterChart()
    #chart.style = 13
    time_data = Reference(data_sheet, min_col=1, min_row=2, max_row=max_idx)
    wind_data = Reference(data_sheet, min_col=2, min_row=2, max_row=max_idx)
    force_data = Reference(data_sheet, min_col=3, min_row=2, max_row=max_idx)
    cal_wind_data = Reference(data_sheet, min_col=4,
                              min_row=2, max_row=max_idx)
    av_force_data = Reference(data_sheet, min_col=6,
                              min_row=2, max_row=max_idx)
    target_force_data = Reference(
        data_sheet, min_col=9, min_row=2, max_row=max_idx)
    drag_area_data = Reference(
        data_sheet, min_col=10, min_row=2, max_row=max_idx)

    wind_series = Series(wind_data, xvalues=time_data, title="Raw Windspeed")
    force_series = Series(force_data, xvalues=time_data, title="Raw Force")
    cal_wind_series = Series(cal_wind_data, xvalues=time_data,
                             title="Calibrated Windspeed")
    av_force_series = Series(av_force_data, xvalues=time_data,
                             title="Averaged Force")
    target_force_series = Series(target_force_data, xvalues=time_data,
                                 title="Target Force")
    drag_area_series = Series(drag_area_data, xvalues=time_data,
                              title="Drag Area")

    chart.append(wind_series)
    chart.append(force_series)
    chart.append(cal_wind_series)
    chart.append(av_force_series)
    chart.append(target_force_series)

    drag_area_chart.append(drag_area_series)

    if DEBUG_MODE_VERBOSE:
        print(data_sheet.cell(max_idx - 1, 1).value)

    chart.x_axis.scaling.max = data_sheet.cell(max_idx - 1, 1).value
    chart.x_axis.scaling.min = 0
    chart.height = 10
    chart.width = 30
    chart.x_axis.title = "Time"

    chart.y_axis.axId = 200

    drag_area_chart.x_axis.scaling.max = data_sheet.cell(max_idx - 1, 1).value
    drag_area_chart.x_axis.scaling.min = 0
    drag_area_chart.height = 10
    drag_area_chart.width = 40
    drag_area_chart.y_axis.title = "Drag Area Axis"
    drag_area_chart.y_axis.majorGridlines = None

    drag_area_chart.y_axis.crosses = "max"

    drag_area_chart += chart
    if output_sheet == None:
        # If there is no separate output sheet specified, the data sheet is the output sheet,
        # with the graph to the side of the data.
        output_sheet = data_sheet
        output_sheet.add_chart(drag_area_chart, 'O2')
    else:
        output_sheet.add_chart(drag_area_chart, 'B2')


def simple_filter(raw_data):  # to be expanded later
    """
    Dummy filter function for sifting through data, currently just returns True.

    Parameters
    ----------
    raw_data: array of float-like str
        The data based on which the filtering will occur.

    Returns
    -------
    bool
        True if data passes the filter criteria, False otherwise.

    """
    return True


def create_averageifs_formula(sheet_name, range_end):
    """
    Parameters
    ----------
    sheet_name: str
        The name of the sheet
    range_end: int
        The end of the range (assuming that the range would start at row 2)
    """
    formula = "=AVERAGEIFS("

    wind_col_letter = get_column_letter(HeaderManager.ANEMOMETER_AVERAGED_COL)
    wind_range_start = wind_col_letter + '2'
    wind_range_end = wind_col_letter + str(range_end)
    wind_range = sheet_name + "!" + wind_range_start + ":" + wind_range_end

    force_col_letter = get_column_letter(HeaderManager.LOAD_CELL_AVERAGED_COL)
    force_range_start = force_col_letter + '2'
    force_range_end = force_col_letter + str(range_end)
    force_range = sheet_name + "!" + force_range_start + ":" + force_range_end

    cdso_col_letter = get_column_letter(HeaderManager.DRAG_AREA_COL)
    cdso_range_start = cdso_col_letter + '2'
    cdso_range_end = cdso_col_letter + str(range_end)
    cdso_range = sheet_name + "!" + cdso_range_start + ":" + cdso_range_end

    formula += (cdso_range)
    formula += ','

    formula += (wind_range)
    formula += ', \">=\"&'
    formula += "WINDSPEED_THRESHOLD, "

    formula += (force_range)
    formula += ', \">=\"&'
    formula += "FORCE_THRESHOLD"

    formula += ")"

    if DEBUG_MODE_VERBOSE:
        print("Avergeifs fnction created: " + formula)

    return formula


def create_meta_analysis(target_workbook, target_worksheet):
    """
    Creates a meta-analysis of the data in a single-file output docuemnt.

    Parameters
    ----------
    target_workbook: openpyxl.Workbook
        The workbook object containing all created worksheets 
    target_worksheet: openpyxl.Worksheet
        The sheet where the meta-analysis should be performed
    """

    all_sheetnames = target_workbook.sheetnames

    row_index = 2

    # Sheets that are excluded from the meta-analysis
    name_blacklist = ['Constants', 'README',
                      'Graphs', "Sheet", "META-ANALYSIS"]
    for idx in range(len(name_blacklist)):
        name_blacklist[idx] = name_blacklist[idx].upper()

    if DEBUG_MODE_VERBOSE:
        print("name_blacklist: ")
        print(name_blacklist)

    for itm in all_sheetnames:
        if not (itm.upper() in name_blacklist):
            target_worksheet.cell(row_index, 1, itm)

            worksheet_dims = target_workbook[itm].calculate_dimension()
            worksheet_end_str = worksheet_dims.split(':')[-1]
            worksheet_end = int(
                "".join(filter(str.isdigit, worksheet_end_str)))

            target_worksheet.cell(row_index, 2, worksheet_end)
            average_formula = create_averageifs_formula(
                quote(target_workbook[itm].title), worksheet_end)
            target_worksheet.cell(row_index, 3, average_formula)

            row_index += 1


def execute_analysis(input_path, output_path,
                     single_file=False,
                     single_workbook=None,
                     sheetname_prefix=None,
                     condensed_version=False):
    """
    Execute a full analysis of a single data set.

    input_path: path-format str
        The path where the csv data is located
    output_path: path-format str
        The path where the excel file should be saved
    single_file: bool
        Whether the analysis should be performed as part of a single output sheet.
        Defaults to False.
    sheetname_prefix: str  
        In the event of a single file output, what the prefix of the sheet would be.
    condensed_version: bool
        In the event of a single file output, whether it should be done in a condensed
        output version. This is mostly used for the google drive output.
    """

    if single_workbook:

        if not condensed_version:
            ws_data = single_workbook.create_sheet(
                title=("Data " + sheetname_prefix))
            ws_graphs = single_workbook.create_sheet(
                title=("Graphs " + sheetname_prefix))
            ws_data.sheet_view.zoomScale = 70
        else:
            ws_data = single_workbook.create_sheet(title=(sheetname_prefix))
            ws_data.sheet_view.zoomScale = 55
    else:
        active_wb = WB()
        ws_data = active_wb.create_sheet(title="Data")
        ws_consts = active_wb.create_sheet(title="Constants")
        ws_graphs = active_wb.create_sheet(title="Graphs")

        Consts().create_constants_table(ws_consts)
        Consts.create_defined_names(active_wb)

    HeaderManager().create_headers(ws_data)

    csvfile = open(input_path + '.csv', newline='')
    csvreader = csv.reader(csvfile, delimiter=',')

    row_idx = 1
    for row in list(csvreader):
        if row_idx > 1:
            if simple_filter(row):  # Data Filtering
                process_sheet_row(row, row_idx, ws_data)
            else:
                row_idx -= 1

        row_idx += 1

    if condensed_version:
        create_graphs(ws_data, None, row_idx)
    else:
        create_graphs(ws_data, ws_graphs, row_idx)

    space_columns(ws_data, 1)

    if not single_workbook:
        space_columns(ws_consts, 10, 1)
        active_wb.save(output_path)


def execute_complete_analysis(config):
    '''
    Executes the full analysis based on the configuration provided to it

    Parameters
    ----------
    config: dict 
        The configuration dictionary. The following fields are included:
         1.  USE_HARD_CODED_PATH: bool : Whether hard-coded paths are used 
                for a single-file analysis
         2.  HARD_CODED_PATH: string : the hard-coded input path
         3.  HARD_CODED_OUTPUT: string: the hard-coded output path
         4.  FILE_INPUT_HITLIST: bool : Whether a input hitlist is used. 
                 Overrides hard-coded paths. (Recommended True)
         5.  HITLIST_PATH: string : the path of the hitlist
         6.  SINGLE_OUTPUT_FILE : bool: Whether the output is in one file or several
         7.  SINGLE_OUTPUT_FILE_PATH: string : What the single output file is to be called
         8.  FILE_SUBDIRECTORY : string : Path of the directory all data inputs and outputs are in
         9.  USE_CUSTOM_SHEETNAMES : bool: If single file output is used, whether custom sheetnames 
                 also used (Recommended True)
         10. SINGLE_OUTPUT_SHEETNAMES_PATH: string : What path the sheetnames list is located at
         11. CONDENSED_EXPORT_VERSION : bool : Whether the single file is condensed or not. This is
                 particularly useful for google drive exports (Recommended True).
         12. SUPPRESS_ALL_PRINTS: bool : Whether all print statements are supressed or not 
    '''

    if not config['SUPPRESS_ALL_PRINTS']:
        print('Analyzer active')

    if config['FILE_INPUT_HITLIST']:  # Performing a full multi-file analysis
        if not config['SUPPRESS_ALL_PRINTS']:
            print("Analyzer operating in input-hitlist mode")

        hitlist_path = config['HITLIST_PATH']
        hitlist = []
        with open(hitlist_path, 'r') as hitlist_file:
            hitlist = list(hitlist_file)

        if len(hitlist) == 0:
            print(
                "No files were specified in the target list. No analysis has been performed.")
            print("Ensure that the correct target file is being passed to the program.")
            print("The current one is " + hitlist_path)
            print("Analyzer finished")
            return

        if config['USE_CUSTOM_SHEETNAMES']:
            sheetnames_path = config['SINGLE_OUTPUT_SHEETNAMES_PATH']
            sheetnames = []
            with open(sheetnames_path, 'r') as sheetnames_file:
                sheetnames = list(sheetnames_file)

            for idx in range(len(sheetnames)):
                sheetnames[idx] = sheetnames[idx].rstrip("\n")
        else:
            # If we are not using custom sheetnames, the sheetnames will be the file names of the
            # files that are to be analyzed. [:] is used to stop issues when full paths are added
            sheetnames = hitlist[:]

        if config['FILE_SUBDIRECTORY']:
            for idx in range(len(hitlist)):
                hitlist[idx] = os.path.join(
                    config['FILE_SUBDIRECTORY'], hitlist[idx])

        # Putting all analysis results into one file
        if config['SINGLE_OUTPUT_FILE']:
            if not config['SUPPRESS_ALL_PRINTS']:
                print('Exporting analysis to single file: '
                      + os.path.join(config['FILE_SUBDIRECTORY'],
                                     config['SINGLE_OUTPUT_FILE_PATH']))

            global_workbook = WB()
            ws_consts = global_workbook.create_sheet(title="Constants")

            from constants import SimplifiedConsts as SConsts
            consts = SConsts(ws_consts, global_workbook)

            # Creating worksheet for performing a summy analysis of all data later
            ws_meta_analysis = global_workbook.create_sheet(
                title="Meta-Analysis")

            space_columns(ws_consts, 12, 1)

            for target, sheetname in zip(hitlist, sheetnames):
                execute_analysis(target[:-1],
                                 output_path=target[:-1] + '___analyzed.xlsx',
                                 single_file=True,
                                 single_workbook=global_workbook,
                                 sheetname_prefix=sheetname,
                                 condensed_version=config['CONDENSED_EXPORT_VERSION'])
                if not config['SUPPRESS_ALL_PRINTS']:
                    print(target[:-1] + " analyzed")

            # Performs meta-analysis to summarize drag area data
            create_meta_analysis(global_workbook, ws_meta_analysis)

            global_workbook.save(os.path.join(
                config['FILE_SUBDIRECTORY'], config['SINGLE_OUTPUT_FILE_PATH']))
            if not config['SUPPRESS_ALL_PRINTS']:
                print('\n Analysis complete')
        else:  # Making an individual analysis file for each input file that was given
            for target in hitlist:
                target_output_path = os.path.join(
                    config['FILE_SUBDIRECTORY'], (target[:-1] + '___analyzed.xlsx'))
                execute_analysis(target[:-1], output_path=target_output_path)
                if not config['SUPPRESS_ALL_PRINTS']:
                    print(target[:-1] + " analyzed")

            if not config['SUPPRESS_ALL_PRINTS']:
                print('\n Analysis complete')
    else:  # Swithcing to simpler mode that analyzes only one file
        if not config['SUPPRESS_ALL_PRINTS']:
            print("Analyzer operating in basic single-file mode")

        if config['USE_HARD_CODED_PATH']:
            itarget = config['HARD_CODED_PATH']
            otarget = config['HARD_CODED_OUTPUT']
        else:
            itarget = input("Enter the input target path")
            otarget = input("Enter the output target path")

        active_wb = WB()
        ws_data = active_wb.create_sheet(title="Data")
        ws_consts = active_wb.create_sheet(title="Constants")
        ws_graphs = active_wb.create_sheet(title="Graphs")

        from constants import SimplifiedConsts as SConsts
        consts = SConsts(ws_consts, active_wb)
        HeaderManager().create_headers(ws_data)

        csvfile = open(itarget, newline='')
        csvreader = csv.reader(csvfile, delimiter=',')

        row_idx = 1
        for row in list(csvreader):
            if row_idx > 1:
                if simple_filter(row):  # Data Filtering
                    process_sheet_row(row, row_idx, ws_data)
                else:
                    row_idx -= 1

            row_idx += 1

        create_graphs(ws_data, ws_graphs, row_idx)

        space_columns(ws_data, 1)
        space_columns(ws_consts, 10, 1)

        active_wb.save(otarget)


if __name__ == "__main__":

    config = {}
    if len(sys.argv) == 2:
        import yaml
        config_path = sys.argv[1]
        with open(config_path) as file:
            config = yaml.load(file, Loader=yaml.FullLoader)
    else:  # Drawing configuration from variables in python script
        # Impored from config_varaibles.py, see file for detailed description
        config['USE_HARD_CODED_PATH'] = USE_HARD_CODED_PATH
        config['HARD_CODED_PATH'] = HARD_CODED_PATH
        config['HARD_CODED_OUTPUT'] = HARD_CODED_OUTPUT
        config['FILE_INPUT_HITLIST'] = FILE_INPUT_HITLIST
        config['HITLIST_PATH'] = HITLIST_PATH
        config['SINGLE_OUTPUT_FILE'] = SINGLE_OUTPUT_FILE
        config['SINGLE_OUTPUT_FILE_PATH'] = SINGLE_OUTPUT_FILE_PATH
        config['FILE_SUBDIRECTORY'] = FILE_SUBDIRECTORY
        config['USE_CUSTOM_SHEETNAMES'] = USE_CUSTOM_SHEETNAMES
        config['SINGLE_OUTPUT_SHEETNAMES_PATH'] = SINGLE_OUTPUT_SHEETNAMES_PATH
        config['CONDENSED_EXPORT_VERSION'] = CONDENSED_EXPORT_VERSION
        config['SUPPRESS_ALL_PRINTS'] = False

    execute_complete_analysis(config)
